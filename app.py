import os
from dotenv import load_dotenv
from pprint import pprint
import json
import datetime

import facebook
import pandas as pd
import numpy as np

from openpyxl import Workbook, load_workbook
load_dotenv() #look in the ".env" file for env vars


def facebookAPI(token, pageInput):

    pageUserName = pageInput + "/feed"

    #connect to Facebook API
    API = facebook.GraphAPI(token)
    
    #get the Facebook data as JSON
    myFieldsList = ["message","created_time","likes.summary(True)","reactions.summary(True)","comments.summary(True)","shares.summary(True)","permalink_url","full_picture"]
    myFields = ",".join(myFieldsList)

    pageJson = API.get_object(pageUserName, fields = myFields)
    
    
    
    #create post class which is used to store data for each Facebook post
    class post:
        def __init__(self, URLparameter, pictureParameter, dateParameter, likesParameter, reactionsParameter, commentsParameter, sharesParameter, captionParameter, commentTextParameter):
            self.URL = URLparameter
            self.picture = pictureParameter
            self.dateUnedited = dateParameter       
            self.likes = likesParameter
            self.reactions = reactionsParameter
            self.comments = commentsParameter
            self.shares = sharesParameter
            self.caption = captionParameter
            self.commentText = commentTextParameter
            
        def getURL(self):
            return self.URL       
    
        def getPicture(self):
            return self.picture  
    
        def getDate(self):
            date = self.dateUnedited.split("T")[0]
            time = self.dateUnedited.split("T")[1]
            hour = time.split(":")[0]
        
            if int(hour) < 4:
                day = int(date.split("-")[2])
                month = date.split("-")[1]
                year = date.split("-")[0]
                
                day -= 1
                date = year + "-" + month + "-" + str(day)
                
            weekdayStrings = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday")
            dateList = date.split("-")
        
            date = datetime.date(int(dateList[0]),int(dateList[1]),int(dateList[2]))
            weekdayInteger = date.weekday()
            
            dictionaryOutput = {"date":date, "weekday":weekdayStrings[weekdayInteger]}
            return dictionaryOutput
    
        def getTime(self):
            time = self.dateUnedited.split("T")[1][0:8]
            hour = time.split(":")[0]
            
            if int(hour) < 4:
                hour = int(time.split(":")[0])
                minute = time.split(":")[1]
                seconds = time.split(":")[2]
                
                hour += 20
                time = str(hour) + ":" + minute + ":" + seconds
                
            else:
                hour = int(time.split(":")[0])
                minute = time.split(":")[1]
                seconds = time.split(":")[2]
                
                hour -= 4
                time = str(hour) + ":" + minute + ":" + seconds
            
            return time
                
        def getLikes(self):
            return self.likes
        
        def getReactions(self):
            return self.reactions
        
        def getComments(self):
            return self.comments
        
        def getShares(self):
            return self.shares
        
        #turn a post's text into a list and remove grammer and non-letter/non-numbers
        #splits a tweet's text into a list of words and a list of hashtags
        def getWords(self):
            postTextString = self.caption
            newPostTextString = ""
        
            #remove non-letters and non-numbers
            for character in postTextString:
                if character.isalnum() or character == " " or character == "#" or character == "'" or character == "-":
                    newPostTextString += character.lower()
            
            postTextList = newPostTextString.split()
            postWordList = []
            postHashtagList = []
            
            for i in range(len(postTextList)):
                if postTextList[i][0] == "#":
                    postHashtagList.append(postTextList[i])
                else:
                    postWordList.append(postTextList[i])
                    
            dictionaryOutput = {"words":postWordList, "hashtags":postHashtagList}
            return dictionaryOutput
        
        def getCaption(self):
            return self.caption
        
        def getCommentText(self):
            return self.commentText
    
    
    #number of posts on page
    numberOfPosts = len(pageJson["data"])
    #list to hold posts
    postList = []
    

    for i in range(numberOfPosts):
        try:
            comments = pageJson["data"][i]["comments"]["summary"]["total_count"]
        except:
            comments = 0
        
        if comments > 0:
            combinedComments = []
            for x in range(comments):
                combinedComments.append(pageJson["data"][i]["comments"]["data"][x]["message"])
        else:
            combinedComments = []
        
        try:
            URL = pageJson["data"][i]["permalink_url"]
        except:
            URL = "null" 
        try:
            picture = pageJson["data"][i]["full_picture"]
        except:
            picture = "null"
        try:
            createdTime = pageJson["data"][i]["created_time"]
        except:
            createdTime = "null"
        try:
            likes = pageJson["data"][i]["likes"]["summary"]["total_count"]
        except:
            likes = 0
        try:
            reactions = pageJson["data"][i]["reactions"]["summary"]["total_count"]
        except:
            reactions = 0
        try:
            shares = pageJson["data"][i]["shares"]["count"]
        except:
            shares = 0
        try:
            caption = pageJson["data"][i]["message"]
        except:
            caption = "null"
            
        postList.append(post(URL, picture, createdTime, likes, reactions, comments, shares, caption, combinedComments))
        
        # Attempt to make a request to the next page of data, if it exists.
        nextPage = pageJson["paging"]["next"]





                             
    
    #for printing all of the post data
    def printPostData(postParameter):
        print("URL: " + str(postParameter.getURL()) + "\n" + 
                "Picture: " + str(postParameter.getPicture()) + "\n" +
                "Date: " + str(postParameter.getDate()["date"]) + "\n" + 
                "Weekday: " + str(postParameter.getDate()["weekday"]) + "\n" + 
                "Time: " + str(postParameter.getTime()) + "\n" + 
                "Likes: " + str(postParameter.getLikes()) + "\n" + 
                "Reactions: " + str(postParameter.getReactions()) + "\n" + 
                "Comments: " + str(postParameter.getComments()) + "\n" + 
                "Shares: " + str(postParameter.getShares()) + "\n" +
                "Caption: " + str(postParameter.getCaption()) + "\n" + 
                "Words: " + str(postParameter.getWords()["words"])+ "\n" + 
                "Hashtags: " + str(postParameter.getWords()["hashtags"])+ "\n" + 
                "Comments Text: " + str(postParameter.getCommentText()) + "\n\n")

    #for printing the post data
    #for i in range(len(postList)):
    #    printPostData(postList[i])

    #create output Excel file
    file_name = pageInput + " Facebook Data.xlsx"
    excelFile = Workbook()
    excelFile.save(file_name)
        
    excelFile = load_workbook(file_name)
    sheet = excelFile.active
    
    #sheet name
    sheet.title = "Posts"
    
    #add headers
    sheet.cell(row = 1, column = 1).value = "Post" 
    sheet.cell(row = 1, column = 2).value = "URL" 
    sheet.cell(row = 1, column = 3).value = "Picture" 
    sheet.cell(row = 1, column = 4).value = "Date" 
    sheet.cell(row = 1, column = 5).value = "Weekday" 
    sheet.cell(row = 1, column = 6).value = "Time" 
    sheet.cell(row = 1, column = 7).value = "Likes" 
    sheet.cell(row = 1, column = 8).value = "Reactions" 
    sheet.cell(row = 1, column = 9).value = "Comments" 
    sheet.cell(row = 1, column = 10).value = "Shares" 
    sheet.cell(row = 1, column = 11).value = "Caption" 
    #sheet.cell(row = 1, column = 12).value = "Words" 
    #sheet.cell(row = 1, column = 13).value = "Hashtags" 
    #sheet.cell(row = 1, column = 14).value = "Comments Text" 

    excelFile.save(file_name)
    
    #adjust column widths
    sheet.column_dimensions['A'].width = 8.26
    sheet.column_dimensions['B'].width = 22.32
    sheet.column_dimensions['C'].width = 13.11
    sheet.column_dimensions['D'].width = 12
    sheet.column_dimensions['E'].width = 9.84
    sheet.column_dimensions['F'].width = 8.26
    sheet.column_dimensions['G'].width = 8.26
    sheet.column_dimensions['H'].width = 8.26
    sheet.column_dimensions['I'].width = 8.26
    sheet.column_dimensions['J'].width = 8.26
    sheet.column_dimensions['K'].width = 20.63
    #sheet.column_dimensions['L'].width = 8.26
    #sheet.column_dimensions['M'].width = 8.26
    #sheet.column_dimensions['N'].width = 15.53
    
    excelFile.save(file_name)
    
    #print post data to the Excel file
    lastRowFree = 2
    
    for i in range(len(postList)):
        sheet.cell(row = lastRowFree, column = 1).value = i + 1
        sheet.cell(row = lastRowFree, column = 2).value = postList[i].getURL()
        sheet.cell(row = lastRowFree, column = 3).value = postList[i].getPicture()
        sheet.cell(row = lastRowFree, column = 4).value = postList[i].getDate()["date"]
        sheet.cell(row = lastRowFree, column = 5).value = postList[i].getDate()["weekday"]
        sheet.cell(row = lastRowFree, column = 6).value = postList[i].getTime() 
        sheet.cell(row = lastRowFree, column = 7).value = postList[i].getLikes()
        sheet.cell(row = lastRowFree, column = 8).value = postList[i].getReactions()
        sheet.cell(row = lastRowFree, column = 9).value = postList[i].getComments()
        sheet.cell(row = lastRowFree, column = 10).value = postList[i].getShares()
        sheet.cell(row = lastRowFree, column = 11).value = postList[i].getCaption()
    
        #format comments text in columns
        #for x in range(postList[i].getComments()):   
        #    sheet.cell(row = lastRowFree, column = 14 + x).value = postList[i].getCommentText()[x]
        
        lastRowFree += 1
        """
        for x in range(len(postList[i].getWords()["words"])):
            sheet.cell(row = lastRowFree, column = 12).value = postList[i].getWords()["words"][x]
            
            sheet.cell(row = lastRowFree, column = 1).value = i + 1
            sheet.cell(row = lastRowFree, column = 2).value = postList[i].getURL()
            sheet.cell(row = lastRowFree, column = 3).value = postList[i].getPicture()
            sheet.cell(row = lastRowFree, column = 4).value = postList[i].getDate()["date"]
            sheet.cell(row = lastRowFree, column = 5).value = postList[i].getDate()["weekday"]
            sheet.cell(row = lastRowFree, column = 6).value = postList[i].getTime() 
            sheet.cell(row = lastRowFree, column = 7).value = postList[i].getLikes()
            sheet.cell(row = lastRowFree, column = 8).value = postList[i].getReactions()
            sheet.cell(row = lastRowFree, column = 9).value = postList[i].getComments()
            sheet.cell(row = lastRowFree, column = 10).value = postList[i].getShares()
            lastRowFree += 1
            
        for x in range(len(postList[i].getWords()["hashtags"])):
            sheet.cell(row = lastRowFree, column = 13).value = postList[i].getWords()["hashtags"][x]
            
            sheet.cell(row = lastRowFree, column = 1).value = i + 1
            sheet.cell(row = lastRowFree, column = 2).value = postList[i].getURL()
            sheet.cell(row = lastRowFree, column = 3).value = postList[i].getPicture()
            sheet.cell(row = lastRowFree, column = 4).value = postList[i].getDate()["date"]
            sheet.cell(row = lastRowFree, column = 5).value = postList[i].getDate()["weekday"]
            sheet.cell(row = lastRowFree, column = 6).value = postList[i].getTime() 
            sheet.cell(row = lastRowFree, column = 7).value = postList[i].getLikes()
            sheet.cell(row = lastRowFree, column = 8).value = postList[i].getReactions()
            sheet.cell(row = lastRowFree, column = 9).value = postList[i].getComments()
            sheet.cell(row = lastRowFree, column = 10).value = postList[i].getShares()
            lastRowFree += 1
            
        for x in range(len(postList[i].getCommentText())):
            sheet.cell(row = lastRowFree, column = 14).value = postList[i].getCommentText()[x]
            
            sheet.cell(row = lastRowFree, column = 1).value = i + 1
            sheet.cell(row = lastRowFree, column = 2).value = postList[i].getURL()
            sheet.cell(row = lastRowFree, column = 3).value = postList[i].getPicture()
            sheet.cell(row = lastRowFree, column = 4).value = postList[i].getDate()["date"]
            sheet.cell(row = lastRowFree, column = 5).value = postList[i].getDate()["weekday"]
            sheet.cell(row = lastRowFree, column = 6).value = postList[i].getTime() 
            sheet.cell(row = lastRowFree, column = 7).value = postList[i].getLikes()
            sheet.cell(row = lastRowFree, column = 8).value = postList[i].getReactions()
            sheet.cell(row = lastRowFree, column = 9).value = postList[i].getComments()
            sheet.cell(row = lastRowFree, column = 10).value = postList[i].getShares()
            lastRowFree += 1
                
    excelFile.save(file_name)
"""


#run the function
FACEBOOK_API_KEY = os.getenv("FACEBOOK_API_KEY")

user_account_search = input("Please enter a Facebook business page: ")

facebookAPI(FACEBOOK_API_KEY, user_account_search)